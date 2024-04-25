from math import log10
from openpyxl import load_workbook, Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
# -------------------
# Read Co-locate.xlsx
# -------------------

wb = load_workbook('Co-locate.xlsx', read_only=True, data_only=True)
ws_Bands = wb['Bands']
ws_OR_Block = wb['Override-blocking']
ws_OR_SpEm = wb['Override-SpEm']

ULfreqStart = ord('H') - ord('A')
CommentCol  = ord('R') - ord('A')
RegCount    = 27
Bands = []
regions = []

# read Bands tab
for row in ws_Bands.rows:
    band = {'ID':           row[0].value.rstrip('_'),
            'MSR':          row[ULfreqStart - 1].value,
            'ULfreqStart':  row[ULfreqStart].value,
            'ULfreqStop':   row[ULfreqStart + 1].value,
            'DLfreqStart':  row[ULfreqStart + 2].value,
            'DLfreqStop':   row[ULfreqStart + 3].value,
            'SensWA':       row[ULfreqStart + 4].value,
            'Comment':      row[CommentCol].value,
            'BlockType':    row[CommentCol + RegCount + 1].value,
            'BlockWA':      row[CommentCol + RegCount + 2].value,
            'BlockAAS':     row[CommentCol + RegCount + 3].value,
            'BlockMR':      row[CommentCol + RegCount + 4].value,
            'BlockLA':      row[CommentCol + RegCount + 5].value,
            'SpEmULWA':     row[CommentCol + RegCount + 6].value,
            'SpEmDLWA':     row[CommentCol + RegCount + 7].value,
            'SpEmULAAS':    row[CommentCol + RegCount + 8].value,
            'SpEmDLAAS':    row[CommentCol + RegCount + 9].value,
            'SpEmULMR':     row[CommentCol + RegCount + 10].value,
            'SpEmDLMR':     row[CommentCol + RegCount + 11].value,
            'SpEmULLA':     row[CommentCol + RegCount + 12].value,
            'SpEmDLLA':     row[CommentCol + RegCount + 13].value}
    if (band['ULfreqStart'] == band['DLfreqStart'] and
        band['ULfreqStop'] == band['DLfreqStop']):
        band['type'] = 'TDD'
    elif band['ULfreqStart'] == None or band['ULfreqStop'] == None:
        band['type'] = 'SDL'
        band['ULfreqStart'] = 'NA'
        band['ULfreqStop'] = 'NA'
    elif band['DLfreqStart'] == None or band['DLfreqStop'] == None:
        band['type'] = 'SUL'
        band['DLfreqStart'] = 'NA'
        band['DLfreqStop'] = 'NA'
    else:
        band['type'] = 'FDD'
    if (isinstance(band['ULfreqStart'], (int, float)) and
        band['ULfreqStart'] > 6000):
        band['FR'] = 2
    elif (isinstance(band['ULfreqStop'], (int, float)) and
        band['ULfreqStop'] > 6000):
        band['FR'] = 2
    elif (isinstance(band['DLfreqStart'], (int, float)) and
        band['DLfreqStart'] > 6000):
        band['FR'] = 2
    elif (isinstance(band['DLfreqStop'], (int, float)) and
        band['DLfreqStop'] > 6000):
        band['FR'] = 2
    else:
        band['FR'] = 1
    Bands.append(band)
    # read regions
    line = []
    noreg = 1
    for col in range(RegCount):
        if row[col + CommentCol + 1].value != None:
            line.append(row[col + CommentCol + 1].value)
            noreg = 0
        else:
            line.append(0)
    line.append(noreg)
    regions.append(line)

# remove headers
Bands.pop(0)
Bands.pop(0)

regions.pop(0)
region_headers = regions.pop(0)
region_headers.pop()
region_headers.append('none')

def read_OR(ws, n = None):
    'read Override data'
    out = []
    for row in ws.rows:
        line = {'table': row[0].value.rstrip('_; ').replace('; ', ';'),
                'row': row[1].value.rstrip('_; ').replace('; ', ';'),
                'f1': row[2].value}
        for col in range(4):
            line[col] = row[col + 3].value
        if n != None:
            line['f2'] = row[7].value
            for col in range(4):
                line[col + 4] = row[col + 8].value
        out.append(line)
    out.pop(0)
    return out

def expand(x):
    for col in ('table', 'row'):
        row = 0
        while row < len(x):
            if ';' in x[row][col]:
                x.insert(row, x[row].copy())
                x[row][col]  = x[row][col].partition(';')[0]
                x[row+1][col]= x[row+1][col].partition(';')[2]
            row +=1
    return x

OR_BlockData = expand(read_OR(ws_OR_Block))
OR_SpEmData = expand(read_OR(ws_OR_SpEm, 'SpEm'))

wb.close()
print('Co-locate.xlsx read')

def LookUp(ID, label):
    if ID.startswith('TDD '):
        ID = ID[4:len(ID)]
    if ID.endswith(', own'):
        ID = ID[0:len(ID) - 5]
    for row in range(len(Bands)):
        if Bands[row]['ID'] == ID:
            return Bands[row][label]
    print(ID, 'not found')

# check for inconsistent OR_SpEmData:
for line in range(len(OR_SpEmData)):
    flag = ''
    if OR_SpEmData[line]['row'] != 'footnote':
        if LookUp(OR_SpEmData[line]['row'], 'type') == 'TDD':
            if OR_SpEmData[line]['f1'] != OR_SpEmData[line]['f2']:
                flag = 'TDD'
            for col in range(4):
                if OR_SpEmData[line][col] != OR_SpEmData[line][col+4]:
                    flag = 'TDD'
        elif LookUp(OR_SpEmData[line]['row'], 'type') == 'SUL':
            if OR_SpEmData[line]['f2'] != None:
                flag = 'SUL'
            for col in range(4):
                if OR_SpEmData[line][col+4] != None:
                    flag = 'SUL'
        elif LookUp(OR_SpEmData[line]['row'], 'type') == 'SDL':
            if OR_SpEmData[line]['f1'] != None:
                flag = 'SDL'
            for col in range(4):
                if OR_SpEmData[line][col] != None:
                    flag = 'SDL'
        if flag != '':
            print(line, 'Override-SpEm for', OR_SpEmData[line]['table'],
                  OR_SpEmData[line]['row'], 'inconsistent', flag)

# ---------------------------
# Add regions and valid bands
# ---------------------------

def Vint(v):
    'int on each element in vector/list'
    out = []
    for n in range(len(v)):
        out.append(int(v[n]))
    return out

def Mcol(M, col):
    'column from matrix (list of lists)'
    out = []
    for n in range(len(M)):
        out.append(M[n][col])
    return out

def Vadd(a, b):
    'vector (list) addition'
    out = []
    for n in range(len(a)):
        out.append(a[n]+b[n])
    return out

def Vsign(v):
    'sign on each element in vector/list'
    out = []
    for n in range(len(v)):
        if v[n] != 0:
            out.append(1)
        else:
            out.append(0)
    return out

for table in range(len(regions)):
    line = ''
    column_vector = [0]*len(regions)
    for col in range(len(region_headers)):
        if regions[table][col] != 0:
            line += region_headers[col] + ', '
            column_vector = Vadd(column_vector, Vint(Mcol(regions, col)))
    Bands[table]['regions'] = line[:len(line)-2] + '.'
    if 'T' in str(Bands[table]['MSR']):
        Bands[table]['valid_bands'] = Vsign(column_vector)

# -----------------------------------------------
# Functions for sorting and identifying sub-bands
# -----------------------------------------------

def CountDig(x):
    "count number of digits directly after 'B' "
    n = 1
    while n < len(x) and x[n].isdecimal():
        n += 1
    return n - 1

def fillstr(x, opt=1):# Used for sorting bands
    "B1A => B0001A, B10A_CN => B0010A@CN. Optional full band suffix @"
    if x[0] != 'B':
        return x
    else:
        out = 'B' + '0' * (4-CountDig(x)) + x[1:].replace('_', '@')
        if CountDig(x) == len(x) - 1 and opt == 1:
            out += '@'
        return out

def mbfill(x): # Used to identify own bands
    "B1-7 => B0001@-B0007@"
    if x[0] != 'B' or not x[1].isdecimal():
        return x
    else:
        n = 1
        out = ''
        while len(x) > 1:
            while n < len(x) and (x[n] not in '-+&'):
                n += 1
            out += fillstr('B' + x[1:n])
            if n < len(x):
                out += x[n]
            x = 'B' + x[n+1:len(x)]
            n = 1
        return out

def bonestr(x): # Used to identify sub-bands
    'TDD B42_JPN, own => B0042'
    if x[0:4] == 'TDD ':
        x = x[4:len(x)]
    if x[0] != 'B':
        return x
    elif CountDig(x) == len(x)-1:
        return fillstr(x, 0)
    elif x[CountDig(x)+1] in '_,-+ @':
        return fillstr(x[:CountDig(x)+1], 0)
    else:
        return fillstr(x)

def sortkey(x):
    if x['DLfreqStart'] != 'NA':
        return 1e6*x['DLfreqStart'] - x['DLfreqStop']
    else:
        return 1e6*x['ULfreqStart'] - x['ULfreqStop']

# -----------------------------------------------------------------
# Add full-band if valid sub-bands completely fill up the full-band
# -----------------------------------------------------------------

for table in range(len(Bands)):
    if 'T' in str(Bands[table]['MSR']):
        for row in range(len(Bands[table]['valid_bands']) - 1):
            if (Bands[table]['valid_bands'][row] == 0 and
                '_' not in Bands[row]['ID'] and
                '-' not in Bands[row]['ID'] and
                '+' not in Bands[row]['ID'] and
                '&' not in Bands[row]['ID'] and
                Bands[row]['MSR'] != None and
                'M' not in Bands[row]['MSR']):
                UL_range = []
                DL_range = []
                line = row + 1
                while line < len(Bands[table]['valid_bands']):
                    if (Bands[table]['valid_bands'][line] == 1 and
                        (bonestr(Bands[row]['ID']) in
                         bonestr(Bands[line]['ID'])) and
                        Bands[row]['type'] == Bands[line]['type'] and
                        Bands[row]['BlockType'] == Bands[line]['BlockType'] and
                        Bands[row]['BlockWA'] == Bands[line]['BlockWA'] and
                        Bands[row]['BlockAAS'] == Bands[line]['BlockAAS'] and
                        Bands[row]['BlockMR'] == Bands[line]['BlockMR'] and
                        Bands[row]['BlockLA'] == Bands[line]['BlockLA'] and
                        Bands[row]['SpEmULWA'] == Bands[line]['SpEmULWA'] and
                        Bands[row]['SpEmDLWA'] == Bands[line]['SpEmDLWA'] and
                        Bands[row]['SpEmULAAS'] == Bands[line]['SpEmULAAS'] and
                        Bands[row]['SpEmDLAAS'] == Bands[line]['SpEmDLAAS'] and
                        Bands[row]['SpEmULMR'] == Bands[line]['SpEmULMR'] and
                        Bands[row]['SpEmDLMR'] == Bands[line]['SpEmDLMR'] and
                        Bands[row]['SpEmULLA'] == Bands[line]['SpEmULLA'] and
                        Bands[row]['SpEmDLLA'] == Bands[line]['SpEmDLLA'] and
                        'DD' in Bands[row]['type'] and
                        (Bands[row]['ULfreqStart'] <=
                         Bands[line]['ULfreqStart']) and 
                        (Bands[row]['ULfreqStop'] >=
                         Bands[line]['ULfreqStop']) and 
                        (Bands[row]['DLfreqStart'] <=
                         Bands[line]['DLfreqStart']) and 
                        (Bands[row]['DLfreqStop'] >=
                         Bands[line]['DLfreqStop']) ):
                        if UL_range == []: # empty
                            UL_range.append(Bands[line]['ULfreqStart'])
                            UL_range.append(Bands[line]['ULfreqStop'])
                            DL_range.append(Bands[line]['DLfreqStart'])
                            DL_range.append(Bands[line]['DLfreqStop'])
                        if (len(UL_range) == 2 and # overlap
                              ((Bands[line]['ULfreqStart'] < UL_range[0] and
                                Bands[line]['ULfreqStop'] >= UL_range[0] and
                                Bands[line]['DLfreqStart'] < DL_range[0] and
                                Bands[line]['DLfreqStop'] >= DL_range[0]) or
                               (Bands[line]['ULfreqStop'] > UL_range[1] and
                                Bands[line]['ULfreqStart'] <= UL_range[1] and
                                Bands[line]['DLfreqStop'] > DL_range[1] and
                                Bands[line]['DLfreqStart'] <= DL_range[1]) or
                               (Bands[line]['ULfreqStart'] >= UL_range[0] and
                                Bands[line]['ULfreqStop'] <= UL_range[1] and
                                Bands[line]['DLfreqStart'] >= DL_range[0] and
                                Bands[line]['DLfreqStop'] <= DL_range[1]))):
                              # expand
                              UL_range[0] = min(UL_range[0],
                                                Bands[line]['ULfreqStart'])
                              UL_range[1] = max(UL_range[1],
                                                Bands[line]['ULfreqStop'])
                              DL_range[0] = min(DL_range[0],
                                                Bands[line]['DLfreqStart'])
                              DL_range[1] = max(DL_range[1],
                                                Bands[line]['DLfreqStop'])
                        if (len(UL_range) == 2 and # full
                            UL_range[0] == Bands[row]['ULfreqStart'] and
                            UL_range[1] == Bands[row]['ULfreqStop'] and
                            DL_range[0] == Bands[row]['DLfreqStart'] and
                            DL_range[1] == Bands[row]['DLfreqStop']):
                            Bands[table]['valid_bands'][row] = 1
                            line = len(Bands[table]['valid_bands']) - 1
                    line += 1

# --------
# Blocking
# --------

def bandIDrule(t, r, b = None):
    "add TDD and own"
    out = Bands[r]['ID']
    if mbfill(Bands[r]['ID']) in mbfill(Bands[t]['ID']):
        out += ', own'
        if b != None and Bands[t]['type'] != 'TDD':
            out += ' ' + b
    if Bands[r]['type'] == 'TDD' and Bands[r]['MSR'] != None:
        out = 'TDD ' + out
    return out

def BlockRule(t, r, typ, v):
    if Bands[r]['type'] == 'SUL':
        return 'NA'
    elif mbfill(Bands[r]['ID']) in mbfill(Bands[t]['ID']): # own band
        if (Bands[t]['type'] == 'TDD'):
            return 'n/a' # TDD (own band)
        elif typ == 'BlockLA':
            return 9  # basis Blocking_LA_from_own_DL
        elif typ == 'BlockMR':
            return 23 # basis Blocking_MR_from_OWN_DL
        else:
            return v
    elif Bands[r][typ] != None:
        return Bands[r][typ]
    elif (typ == 'BlockMR' and
          Bands[r]['MSR'] != None and
          Bands[r]['DLfreqStart'] < 1000):
        if Bands[t]['type'] != 'SUL':
            if Bands[t]['DLfreqStart'] < 1000:
                return 18 # basis Blocking_MR_LB_from_MR_LB
            else:
                return v
        elif Bands[t]['ULfreqStart'] < 1000:
            return 18  # basis Blocking_MR_LB_from_MR_LB
        else:
            return v
    else:
        return v

def BlockRuleType(x):
    if x == None:
        return 'WB'
    else:
        return x

# build blocking tables
for table in range(len(Bands)):
    if 'T' in str(Bands[table]['MSR']):
        tab = []
        for row in range(len(Bands)):
            if (Bands[table]['valid_bands'][row] == 1 and
                Bands[row]['type'] != 'SUL' and not
                 (Bands[row]['BlockWA'] == 'NA' and
                  Bands[row]['BlockMR'] == 'NA' and
                  Bands[row]['BlockAAS'] == 'NA' and
                  Bands[row]['BlockLA'] == 'NA')):
                tab.append({'ID': bandIDrule(table, row, 'DL'),
                            'DLfreqStart':  Bands[row]['DLfreqStart'],
                            'DLfreqStop':   Bands[row]['DLfreqStop'],
                            'BlockWA':      BlockRule(table, row,
                                                      'BlockWA', 22),
                                            # basis Blocking_WA
                            'BlockAAS':     BlockRule(table, row,
                                                      'BlockAAS', 12),
                            'BlockMR':      BlockRule(table, row,
                                                      'BlockMR', 8),
                                            # basis Blocking_MR
                            'BlockLA':      BlockRule(table, row,
                                                      'BlockLA', -6),
                                            # basis Blocking_LA_3GPP
                            'BlockType':    BlockRuleType(Bands[row]
                                                          ['BlockType']),
                            'FR':           Bands[row]['FR']})
                if (mbfill(Bands[row]['ID']) in mbfill(Bands[table]['ID']) and
                    Bands[table]['type'] == 'FDD' and
                    Bands[row]['type'] != 'SDL'):
                    # add own UL
                    tab.insert(0, {'ID': bandIDrule(table, row, 'UL'),
                                   'DLfreqStart':   Bands[row]['ULfreqStart'],
                                   'DLfreqStop':    Bands[row]['ULfreqStop'],
                                   'BlockWA': 'own UL',
                                   'BlockAAS': 'own UL',
                                   'BlockMR': 'own UL',
                                   'BlockLA': 'own UL',
                                   'BlockType': '',
                                   'FR':            Bands[row]['FR']})
        Bands[table]['BlockT'] = tab
    else:
        Bands[table]['BlockT'] = None


for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        # remove FR2 rows from FR1 tables
        row = 0
        while row < len(Bands[table]['BlockT']):
            if (Bands[table]['FR'] == 1 and
                Bands[table]['BlockT'][row]['FR'] == 2):
                Bands[table]['BlockT'].pop(row)
                row -= 1
            row += 1
        # remove sub-bands in blocking tables
        row = 1
        while row < len(Bands[table]['BlockT'])-2:
            if ((bonestr(Bands[table]['BlockT'][row - 1]['ID']) in
                 bonestr(Bands[table]['BlockT'][row]['ID'])) and
                (Bands[table]['BlockT'][row - 1]['DLfreqStart'] <=
                 Bands[table]['BlockT'][row]['DLfreqStart']) and
                (Bands[table]['BlockT'][row - 1]['DLfreqStop'] >=
                 Bands[table]['BlockT'][row]['DLfreqStop']) and
                ((Bands[table]['BlockT'][row - 1]['BlockWA'] ==
                  Bands[table]['BlockT'][row]['BlockWA'] and
                  Bands[table]['BlockT'][row - 1]['BlockAAS'] ==
                  Bands[table]['BlockT'][row]['BlockAAS'] and
                  Bands[table]['BlockT'][row - 1]['BlockMR'] ==
                  Bands[table]['BlockT'][row]['BlockMR'] and
                  Bands[table]['BlockT'][row - 1]['BlockLA'] ==
                  Bands[table]['BlockT'][row]['BlockLA'] and
                  Bands[table]['BlockT'][row - 1]['BlockType'] ==
                  Bands[table]['BlockT'][row]['BlockType']) or
                 ('own' in Bands[table]['BlockT'][row - 1]['ID']))):
                # remove sub-band
                Bands[table]['BlockT'].pop(row)
            elif ((bonestr(Bands[table]['BlockT'][row - 1]['ID']) in
                   bonestr(Bands[table]['BlockT'][row]['ID'])) and
                  (Bands[table]['BlockT'][row - 1]['DLfreqStart'] <=
                   Bands[table]['BlockT'][row]['DLfreqStart']) and
                  (Bands[table]['BlockT'][row - 1]['DLfreqStop'] >=
                   Bands[table]['BlockT'][row]['DLfreqStop'])):
                # swap rows and continue
                temp = Bands[table]['BlockT'][row-1]
                Bands[table]['BlockT'][row-1] = Bands[table]['BlockT'][row]
                Bands[table]['BlockT'][row] = temp
                row += 1
            else:
                row += 1
        # sort blocking tables
        Bands[table]['BlockT'].sort(key = sortkey)
        # 'NA' for bands in own UL
        if (('M' not in Bands[table]['MSR'] or
             '-' not in Bands[table]['ID']) and
            str(Bands[table]['type']) in ['FDD', 'SUL', 'TDD']):
            for row in range(len(Bands[table]['BlockT'])):
                if (', own' not in Bands[table]['BlockT'][row]['ID'] and
                    Bands[table]['ULfreqStart'] <=
                    Bands[table]['BlockT'][row]['DLfreqStart'] and
                    Bands[table]['ULfreqStop'] >=
                    Bands[table]['BlockT'][row]['DLfreqStop']):
                    #print('Setting NA in', Bands[table]['ID'],
                    #      Bands[table]['BlockT'][row]['ID'])
                    Bands[table]['BlockT'][row]['BlockWA']  = 'NA'
                    Bands[table]['BlockT'][row]['BlockAAS'] = 'NA'
                    Bands[table]['BlockT'][row]['BlockMR']  = 'NA'
                    Bands[table]['BlockT'][row]['BlockLA']  = 'NA'
                elif (', own' not in Bands[table]['BlockT'][row]['ID'] and
                      Bands[table]['type'] == 'TDD' and
                      Bands[table]['BlockT'][row]['ID'][:3] == 'TDD' and
                      ((Bands[table]['ULfreqStart'] >=
                        Bands[table]['BlockT'][row]['DLfreqStart'] and
                        Bands[table]['ULfreqStart'] <=
                        Bands[table]['BlockT'][row]['DLfreqStop']) or
                       (Bands[table]['ULfreqStop'] >=
                        Bands[table]['BlockT'][row]['DLfreqStart'] and
                        Bands[table]['ULfreqStop'] <=
                        Bands[table]['BlockT'][row]['DLfreqStop']))):
                    #print('Setting NA for TDD partial in', Bands[table]['ID'],
                    #      Bands[table]['BlockT'][row]['ID'])
                    Bands[table]['BlockT'][row]['BlockWA']  = 'NA'
                    Bands[table]['BlockT'][row]['BlockAAS'] = 'NA'
                    Bands[table]['BlockT'][row]['BlockMR']  = 'NA'
                    Bands[table]['BlockT'][row]['BlockLA']  = 'NA'
        # over ride blocking tables
        row = 0
        while row < len(Bands[table]['BlockT']):
            pointer = 0
            while pointer < len(OR_BlockData):
                if (Bands[table]['ID'] == OR_BlockData[pointer]['table'] and
                    Bands[table]['BlockT'][row]['ID'] == 
                    OR_BlockData[pointer]['row']):
                    flagnewline = 0
                    if (OR_BlockData[pointer]['f1'] != None ):
                        if (OR_BlockData[pointer]['f1'] <
                            Bands[table]['BlockT'][row]['DLfreqStop']):
                            # insert new row
                            Bands[table]['BlockT'].\
                            insert(row, Bands[table]['BlockT'][row].copy())
                            row += 1
                            flagnewline = 1
                        else:
                            flagnewline = 0
                        if (OR_BlockData[pointer]['f1'] >=
                            Bands[table]['BlockT'][row]['DLfreqStart'] and
                            OR_BlockData[pointer]['f1'] <
                            Bands[table]['BlockT'][row]['DLfreqStop']):
                            # frequency break point
                            Bands[table]['BlockT'][row - 1]['DLfreqStop'] = \
                            OR_BlockData[pointer]['f1']
                            Bands[table]['BlockT'][row]['DLfreqStart'] = \
                            OR_BlockData[pointer]['f1']
                    for n in range(4):
                        if OR_BlockData[pointer][n] != None:
                            if (Bands[table]['ID'] !=
                                OR_BlockData[pointer]['table'] or
                                Bands[table]['BlockT'][row-flagnewline]['ID']
                                != OR_BlockData[pointer]['row']):
                                print(OR_BlockData[pointer]['table'],
                                      OR_BlockData[pointer]['row'],
                                      'mismatch b')
                            if Bands[table]['BlockT'][row-flagnewline]\
                               [['BlockWA', 'BlockAAS',
                                 'BlockMR', 'BlockLA'][n]] == 'NA':
                                print('For', Bands[table]['ID'],
                                      Bands[table]['BlockT'][row]['ID'],
                                      'owerwrite NA with',
                                      OR_BlockData[pointer][n])
                            Bands[table]['BlockT']\
                            [row - flagnewline][['BlockWA',
                                                 'BlockAAS',
                                                 'BlockMR',
                                                 'BlockLA'][n]] = \
                            OR_BlockData[pointer][n]
                    if (OR_BlockData[pointer]['f1'] == None or
                        OR_BlockData[pointer]['f1'] >=
                        Bands[table]['BlockT'][row]['DLfreqStop']):
                        OR_BlockData.pop(pointer)
                        pointer = len(OR_BlockData)
                    else:
                        OR_BlockData.pop(pointer)
                        pointer -= 1
                pointer += 1
            row += 1
print('Override-Blocking not inserted: ', OR_BlockData) # remains


# add selectivity


def isnumber(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def Sel(table, row, tableID = None):
    REFSENS = Bands[table]['SensWA']
    if isinstance(REFSENS, list): # multi-band
        output = []
        for i in range(len(REFSENS)):
            Bands[table]['SensWA'] = REFSENS[i][1]
            output.append([REFSENS[i][0], Sel(table, row, REFSENS[i][0])])
            Bands[table]['SensWA'] = REFSENS
        allsame = True
        for i in range(len(output)-1):
            if output[i+1][1] != output[0][1]:
                allsame = False
        if allsame:
            return output[0][1]
        outputstring = ''
        for i in range(len(output)):
            outputstring += ', ' + output[i][0] + ': ' + str(output[i][1])
        return outputstring[2:]
    I = Bands[table]['BlockT'][row]['BlockWA']
    SD = 0.4
    SDadd = 0 # additional SD (if relative to 3GPP ref sens)
    SNR = -1.1 # LTE5
    if ('CW' in str(Bands[table]['BlockType']) and
        isinstance(REFSENS, (int, float))):
        REFSENS -= SNR
        SNR = 4.4 # GSM
        REFSENS += SNR
        REFSENS += 10*log10(0.24/4.5) # BW GSM vs BW LTE5
    if isinstance(I, str):
        if I in ('NA', 'n/a', 'own UL'):
            return I
        elif (tableID != None and
              ' if ' in I and ' off' in I and ', else ' in I):
            if tableID in I:
                Bands[table]['BlockT'][row]['BlockWA'] = \
                I.partition(', else ')[2]
            else:
                Bands[table]['BlockT'][row]['BlockWA'] = \
                I.partition(' if ')[0]
            if isnumber(Bands[table]['BlockT'][row]['BlockWA']):
                Bands[table]['BlockT'][row]['BlockWA'] = \
                float(Bands[table]['BlockT'][row]['BlockWA'])
            sel = Sel(table, row)
            Bands[table]['BlockT'][row]['BlockWA'] = I
            return sel
        elif ' and ' in I and ' if ' not in I:
            # two cases to check worst case for
            Bands[table]['BlockT'][row]['BlockWA'] = I.partition(' and ')[0]
            sel1 = Sel(table, row)
            Bands[table]['BlockT'][row]['BlockWA'] = I.partition(' and ')[2]
            sel2 = Sel(table, row)
            Bands[table]['BlockT'][row]['BlockWA'] = I
            return min(sel1, sel2)
        
        elif (I.endswith(' rel 3GPP ref sens') and
              isinstance(REFSENS, (int, float))):
            SDadd = -101.5 - REFSENS
            I = I[:len(I)-18]
        if (' @ ' in I and I.endswith(' dB SD') and
            isnumber(I.partition(' @ ')[0]) and
            isnumber(I.partition(' @ ')[2].replace(' dB SD', ''))):
            SD = float(I.partition(' @ ')[2].replace(' dB SD', ''))
            I = float(I.partition(' @ ')[0])
    if isinstance(I, (int, float)) and isinstance(REFSENS, (int, float)):
        return round(REFSENS-I-SNR+10*log10(10**((SD+SDadd)/10)-1), 1)
    else:
        return '???'

for table in range(len(Bands)):
    # Sensitivity for all bands in multi-band:
    if (Bands[table]['MSR'] != None and 'M' in Bands[table]['MSR'] and
        Bands[table]['SensWA'] == None):
        temp = []
        for row in range(len(Bands)):
            if (row != table and
                Bands[row]['MSR'] != None and 'M' not in Bands[row]['MSR'] and
                '&' not in Bands[row]['ID'] and
                mbfill(Bands[row]['ID']) in mbfill(Bands[table]['ID'])):
                temp.append([Bands[row]['ID'], Bands[row]['SensWA']])
        temp.sort(key=lambda x: fillstr(x[0]))
        Bands[table]['SensWA'] = temp
        # print(Bands[table]['ID'], Bands[table]['SensWA'])
    # Calculate selectivity:
    if Bands[table]['BlockT'] != None:
        for row in range(len(Bands[table]['BlockT'])):
            Bands[table]['BlockT'][row]['SelWA'] = Sel(table, row)
            # Bands[table]['BlockT'][row]['SelAAS'] = '???'
            # Bands[table]['BlockT'][row]['SelMR'] = '???'
            # Bands[table]['BlockT'][row]['SelLA'] = '???'


# ----
# SpEm
# ----

def eirp2arp(r, typ):
    # Convert to conducted via antenna gain and feeder loss
    if Bands[r]['type'] == 'SUL':
        freq = Bands[r]['ULfreqStop']
    elif Bands[r]['type'] == 'SDL':
        freq = Bands[r]['DLfreqStop']
    else:
        freq = max(Bands[r]['ULfreqStop'], Bands[r]['DLfreqStop'])
    if typ[-2:] == 'WA':
        Bands[r][typ] = Bands[r][typ][:-5]
        if freq <= 1550:
            gain = 16
        elif freq <= 3100:
            gain = 18
        else:
            gain = 20
        if '/' in Bands[r][typ]:
            Bands[r][typ] = (str(eval(Bands[r][typ].partition('/')[0])-gain) +
                             '/' + Bands[r][typ].partition('/')[2])
        else:
            Bands[r][typ] = eval(Bands[r][typ]) - gain
    elif typ[-2:] == 'AS':
        Bands[r][typ] = Bands[r][typ][:-5]
        if '/' in Bands[r][typ]:
            Bands[r][typ] = (str(eval(Bands[r][typ].partition('/')[0]) - 20) +
                             '/' + Bands[r][typ].partition('/')[2])
        else:
            Bands[r][typ] = eval(Bands[r][typ]) - 20
    elif typ[-2:] == 'MR':
        Bands[r][typ] = Bands[r][typ][:-5]
        if freq <= 3100 and '/' in Bands[r][typ]:
            Bands[r][typ] = (str(eval(Bands[r][typ].partition('/')[0]) - 10) +
                             '/' + Bands[r][typ].partition('/')[2])
        elif freq <= 3100:
            Bands[r][typ] = eval(Bands[r][typ]) - 10
        elif freq <= 4200 and '/' in Bands[r][typ]:
            Bands[r][typ] = (str(eval(Bands[r][typ].partition('/')[0]) - 16) +
                             '/' + Bands[r][typ].partition('/')[2] +
                             ' if 2T2R, else ' +
                             str(eval(Bands[r][typ].partition('/')[0]) - 13) +
                             '/' + Bands[r][typ].partition('/')[2])
        elif freq <= 4200:
            Bands[r][typ] = (str(eval(Bands[r][typ]) - 16) +
                             ' if 2T2R, else ' +
                             str(eval(Bands[r][typ]) - 13))
        else:
            print('>4200 MHz & 2T2R vs 4T4R')
            breakpoint()
    elif typ[-2:] == 'LA':
        Bands[r][typ] = Bands[r][typ][:-5]
        if '/' not in Bands[r][typ]:
            Bands[r][typ] = eval(Bands[r][typ]) # 0 dB

def SpEmULrule(t, r, typ, v):
    if Bands[r]['type'] == 'SDL':
        return 'NA'
    elif ((mbfill(Bands[r]['ID']) in mbfill(Bands[t]['ID'])) and
          Bands[t]['type'] != 'TDD'):
        return 'own UL (ch 6.6.1.2)'
    elif ((mbfill(Bands[r]['ID']) in mbfill(Bands[t]['ID'])) and
          Bands[t]['type'] == 'TDD'):
        return 'own'
    elif Bands[r][typ] != None:
        if isinstance(Bands[r][typ], str) and Bands[r][typ][-5:] == ' EIRP':
            eirp2arp(r, typ)
        return Bands[r][typ]
    elif (typ == 'SpEmULMR' and
          Bands[r]['MSR'] != None and
          Bands[r]['ULfreqStart'] < 1000):
        if Bands[t]['type'] != 'SUL':
            if Bands[t]['DLfreqStart'] < 1000:
                return -101 # basis SpEmUL_MR_LB_to_LB
            else:
                return v
        elif Bands[t]['ULfreqStart'] < 1000:
            return -101  # basis SpEmUL_MR_LB_to_LB
        else:
            return v
    else:
        return v

def SpEmDLrule(t, r, typ, v):
    if Bands[r]['type'] == 'TDD' and typ == 'SpEmDLWA':
        return SpEmULrule(t, r, 'SpEmULWA', -101)
    elif Bands[r]['type'] == 'TDD' and typ == 'SpEmDLAAS':
        return SpEmULrule(t, r, 'SpEmULAAS', -91)
    elif Bands[r]['type'] == 'TDD' and typ == 'SpEmDLMR':
        return SpEmULrule(t, r, 'SpEmULMR', -91)
    elif Bands[r]['type'] == 'TDD' and typ == 'SpEmDLLA':
        return SpEmULrule(t, r, 'SpEmULLA', -88)
    elif Bands[r][typ] != None:
        if isinstance(Bands[r][typ], str) and Bands[r][typ][-5:] == ' EIRP':
            eirp2arp(r, typ)
        return Bands[r][typ]
    elif ((mbfill(Bands[r]['ID']) in mbfill(Bands[t]['ID'])) or
        Bands[r]['type'] == 'SUL'):
        return 'NA'
    elif Bands[r]['MSR'] != None or typ == 'SpEmDLAAS':
        return v
    else:
        return -57
        # basis SpEmDL_WA_vs_Other, SpEmDL_MR_to_other, SpEmDL_LA_to_other

# build SpEm tables
for table in range(len(Bands)):
    if 'T' in str(Bands[table]['MSR']):
        tab = []
        for row in range(len(Bands)):
            if Bands[table]['valid_bands'][row] == 1:
                tab.append({'ID': bandIDrule(table, row),
                            'ULfreqStart':  Bands[row]['ULfreqStart'],
                            'ULfreqStop':   Bands[row]['ULfreqStop'],
                            'SpEmULWA':
                            SpEmULrule(table, row, 'SpEmULWA', -101),
                                    # basis SpEmUL_WA_to_WA
                            'SpEmULAAS':
                            SpEmULrule(table, row, 'SpEmULAAS', -91),
                            'SpEmULMR':
                            SpEmULrule(table, row, 'SpEmULMR', -91),
                                    # basis SpEmUL_MR
                            'SpEmULLA':
                            SpEmULrule(table, row, 'SpEmULLA', -88),
                                    # basis SpEmUL_LA
                            'DLfreqStart': Bands[row]['DLfreqStart'],
                            'DLfreqStop':  Bands[row]['DLfreqStop'],
                            'SpEmDLWA':
                            SpEmDLrule(table, row, 'SpEmDLWA', -62),
                                    # basis SpEmDL_WA
                            'SpEmDLAAS':
                            SpEmDLrule(table, row, 'SpEmDLAAS', -53),
                            'SpEmDLMR':
                            SpEmDLrule(table, row, 'SpEmDLMR', -79),
                                    # basis SpEmDL_MR
                            'SpEmDLLA':
                            SpEmDLrule(table, row, 'SpEmDLLA', -87),
                                    # basis SpEmDL_LA
                            'FR':           Bands[row]['FR']})
                                    
        Bands[table]['SpEmT'] = tab
        del Bands[table]['valid_bands']
    else:
        Bands[table]['SpEmT'] = None

for table in range(len(Bands)):
    if Bands[table]['SpEmT'] != None:
        # remove FR2 rows from FR1 table, except FPU*
        row = 0
        while row < len(Bands[table]['SpEmT']):
            if (Bands[table]['FR'] == 1 and
                Bands[table]['SpEmT'][row]['FR'] == 2
                # and Bands[table]['SpEmT'][row]['ID'][:3] != 'FPU'
                ):
                Bands[table]['SpEmT'].pop(row)
                row -= 1
            row += 1
        # remove 'NA' lines in SpEm tables
        row = 0
        while row < len(Bands[table]['SpEmT']):
            if (('own' not in Bands[table]['SpEmT'][row]['ID']) and
                Bands[table]['SpEmT'][row]['SpEmULWA'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmULAAS'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmULMR'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmULLA'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmDLWA'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmDLAAS'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmDLMR'] == 'NA' and
                Bands[table]['SpEmT'][row]['SpEmDLLA'] == 'NA'):
                # remove line
                Bands[table]['SpEmT'].pop(row)
                row -= 1
            row += 1
        # remove sub-bands in SpEm tables
        row = 1
        while row < len(Bands[table]['SpEmT'])-2:
            if ((bonestr(Bands[table]['SpEmT'][row - 1]['ID']) in
                 bonestr(Bands[table]['SpEmT'][row]['ID'])) and
                (Bands[table]['SpEmT'][row - 1]['DLfreqStart'] <=
                 Bands[table]['SpEmT'][row]['DLfreqStart']) and
                (Bands[table]['SpEmT'][row - 1]['DLfreqStop'] >=
                 Bands[table]['SpEmT'][row]['DLfreqStop']) and
                (Bands[table]['SpEmT'][row - 1]['ULfreqStart'] <=
                 Bands[table]['SpEmT'][row]['ULfreqStart']) and
                (Bands[table]['SpEmT'][row - 1]['ULfreqStop'] >=
                 Bands[table]['SpEmT'][row]['ULfreqStop']) and
                ((Bands[table]['SpEmT'][row - 1]['SpEmULWA'] ==
                  Bands[table]['SpEmT'][row]['SpEmULWA'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmULAAS'] ==
                  Bands[table]['SpEmT'][row]['SpEmULAAS'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmULMR'] ==
                  Bands[table]['SpEmT'][row]['SpEmULMR'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmULLA'] ==
                  Bands[table]['SpEmT'][row]['SpEmULLA'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmDLWA'] ==
                  Bands[table]['SpEmT'][row]['SpEmDLWA'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmDLAAS'] ==
                  Bands[table]['SpEmT'][row]['SpEmDLAAS'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmDLMR'] ==
                  Bands[table]['SpEmT'][row]['SpEmDLMR'] and
                  Bands[table]['SpEmT'][row - 1]['SpEmDLLA'] ==
                  Bands[table]['SpEmT'][row]['SpEmDLLA']) or
                 ('own' in Bands[table]['SpEmT'][row - 1]['ID']))):
                # remove sub-band
                Bands[table]['SpEmT'].pop(row)
            elif ((bonestr(Bands[table]['SpEmT'][row - 1]['ID']) in
                   bonestr(Bands[table]['SpEmT'][row]['ID'])) and
                  (Bands[table]['SpEmT'][row - 1]['DLfreqStart'] <=
                   Bands[table]['SpEmT'][row]['DLfreqStart']) and
                  (Bands[table]['SpEmT'][row - 1]['DLfreqStop'] >=
                   Bands[table]['SpEmT'][row]['DLfreqStop']) and
                  (Bands[table]['SpEmT'][row - 1]['ULfreqStart'] <=
                   Bands[table]['SpEmT'][row]['ULfreqStart']) and
                  (Bands[table]['SpEmT'][row - 1]['ULfreqStop'] >=
                   Bands[table]['SpEmT'][row]['ULfreqStop'])):
                # swap rows and continue
                temp = Bands[table]['SpEmT'][row-1]
                Bands[table]['SpEmT'][row-1] = Bands[table]['SpEmT'][row]
                Bands[table]['SpEmT'][row] = temp
                row += 1
            else:
                row += 1
        # sort SpEm tables
        Bands[table]['SpEmT'].sort(key = sortkey)
        # 'NA' for bands in own DL
        if (('M' not in Bands[table]['MSR'] or
             '-' not in Bands[table]['ID']) and
            str(Bands[table]['type']) in ['FDD', 'SDL', 'TDD']):
            for row in range(len(Bands[table]['SpEmT'])):
                for ULDL in ['UL', 'DL']:
                    if (', own' not in Bands[table]['SpEmT'][row]['ID'] and
                        isinstance(Bands[table]['SpEmT'][row][ULDL +
                                                              'freqStart'],
                                   (int, float)) and
                        isinstance(Bands[table]['SpEmT'][row][ULDL +
                                                              'freqStop'],
                                   (int, float)) and
                        Bands[table]['DLfreqStart'] <=
                        Bands[table]['SpEmT'][row][ULDL +
                                                   'freqStart'] and
                        Bands[table]['DLfreqStop'] >=
                        Bands[table]['SpEmT'][row][ULDL +
                                                   'freqStop']):
                        #print('Setting NA in', Bands[table]['ID'],
                        #      Bands[table]['SpEmT'][row]['ID'], ULDL)
                        for cl in ['WA', 'AAS', 'MR', 'LA']:
                            Bands[table]['SpEmT'][row]['SpEm' +
                                                       ULDL + cl] = 'NA'
                    elif (', own' not in Bands[table]['SpEmT'][row]['ID'] and
                          Bands[table]['type'] == 'TDD' and
                          Bands[table]['SpEmT'][row]['ID'][:3] == 'TDD' and
                          ((Bands[table]['DLfreqStart'] >=
                            Bands[table]['SpEmT'][row][ULDL + 'freqStart'] and
                            Bands[table]['DLfreqStart'] <=
                            Bands[table]['SpEmT'][row][ULDL + 'freqStop']) or
                           (Bands[table]['DLfreqStop'] >=
                            Bands[table]['SpEmT'][row][ULDL + 'freqStart'] and
                            Bands[table]['DLfreqStop'] <=
                            Bands[table]['SpEmT'][row][ULDL + 'freqStop']))):
                        #print('Setting NA for TDD partial in',
                        #      Bands[table]['ID'],
                        #      Bands[table]['SpEmT'][row]['ID'], ULDL)
                        for cl in ['WA', 'AAS', 'MR', 'LA']:
                            Bands[table]['SpEmT'][row]['SpEm' +
                                                       ULDL + cl] = 'NA'
                # NA for DL in FDD/SDL sub-band's full band
                if (bonestr(Bands[table]['SpEmT'][row]['ID']) in
                    bonestr(Bands[table]['ID']) and
                    'own' not in Bands[table]['SpEmT'][row]['ID']):
                    if (Bands[table]['type'] in ('FDD', 'SDL') and
                        Bands[table]['SpEmT'][row]['DLfreqStart'] <=
                        Bands[table]['DLfreqStart'] and
                        Bands[table]['SpEmT'][row]['DLfreqStop'] >=
                        Bands[table]['DLfreqStop']):
                        Bands[table]['SpEmT'][row]['SpEmDLWA'] = 'NA'
                        Bands[table]['SpEmT'][row]['SpEmDLAAS'] = 'NA'
                        Bands[table]['SpEmT'][row]['SpEmDLMR'] = 'NA'
                        Bands[table]['SpEmT'][row]['SpEmDLLA'] = 'NA'
        # over ride
        row = 0
        while row < len(Bands[table]['SpEmT']):
            pointer = 0
            while pointer < len(OR_SpEmData):
                if (Bands[table]['ID'] == OR_SpEmData[pointer]['table'] and
                    Bands[table]['SpEmT'][row]['ID'] ==
                    OR_SpEmData[pointer]['row']):
                    if (OR_SpEmData[pointer]['f1'] != None or
                        OR_SpEmData[pointer]['f2'] != None ):
                        if ((OR_SpEmData[pointer]['f1'] != None and
                             OR_SpEmData[pointer]['f1'] <
                             Bands[table]['SpEmT'][row]['ULfreqStop']) or
                            (OR_SpEmData[pointer]['f2'] != None and
                             OR_SpEmData[pointer]['f2'] <
                             Bands[table]['SpEmT'][row]['DLfreqStop'])):
                            # insert new row
                            Bands[table]['SpEmT'].\
                            insert(row, Bands[table]['SpEmT'][row].copy())
                            row += 1
                            flagnewline = 1
                        else:
                            flagnewline = 0
                        if ((OR_SpEmData[pointer]['f1'] != None and
                             OR_SpEmData[pointer]['f1'] >
                             Bands[table]['SpEmT'][row]['ULfreqStart'] and
                             OR_SpEmData[pointer]['f1'] <
                             Bands[table]['SpEmT'][row]['ULfreqStop']) or
                            (OR_SpEmData[pointer]['f2'] != None and
                             OR_SpEmData[pointer]['f2'] >
                             Bands[table]['SpEmT'][row]['DLfreqStart'] and
                             OR_SpEmData[pointer]['f2'] <
                             Bands[table]['SpEmT'][row]['DLfreqStop']) ):
                            # frequency break point
                            if OR_SpEmData[pointer]['f1'] != None:
                                Bands[table]['SpEmT'][row - 1]['ULfreqStop'] =\
                                OR_SpEmData[pointer]['f1']
                                Bands[table]['SpEmT'][row]['ULfreqStart'] =\
                                OR_SpEmData[pointer]['f1']
                            if OR_SpEmData[pointer]['f2'] != None:
                                Bands[table]['SpEmT'][row - 1]['DLfreqStop'] =\
                                OR_SpEmData[pointer]['f2']
                                Bands[table]['SpEmT'][row]['DLfreqStart'] =\
                                OR_SpEmData[pointer]['f2']
                    for n in range(8):
                        if OR_SpEmData[pointer][n] != None:
                            if (Bands[table]['ID'] !=
                                OR_SpEmData[pointer]['table'] or
                                Bands[table]['SpEmT'][row-flagnewline]['ID']
                                != OR_SpEmData[pointer]['row']):
                                print(OR_SpEmData[pointer]['table'],
                                      OR_SpEmData[pointer]['row'],
                                      'mismatch t')
                            if (Bands[table]['SpEmT'][row-flagnewline]\
                                [['SpEmULWA', 'SpEmULAAS', 'SpEmULMR',
                                  'SpEmULLA', 'SpEmDLWA', 'SpEmDLAAS',
                                  'SpEmDLMR', 'SpEmDLLA'][n]] == 'NA' and
                                'if ' not in str(OR_SpEmData[pointer][n])):
                                print('For', Bands[table]['ID'],
                                      Bands[table]['SpEmT'][row]['ID'],
                                      ['SpEmULWA', 'SpEmULAAS', 'SpEmULMR',
                                       'SpEmULLA', 'SpEmDLWA', 'SpEmDLAAS',
                                       'SpEmDLMR', 'SpEmDLLA'][n],
                                      'owerwrite NA with',
                                      OR_SpEmData[pointer][n])
                            Bands[table]['SpEmT'][row - flagnewline]\
                            [['SpEmULWA',
                              'SpEmULAAS',
                              'SpEmULMR',
                              'SpEmULLA',
                              'SpEmDLWA',
                              'SpEmDLAAS',
                              'SpEmDLMR',
                              'SpEmDLLA'][n]] = OR_SpEmData[pointer][n]
                    if ((OR_SpEmData[pointer]['f1'] == None and
                         OR_SpEmData[pointer]['f2'] == None) or
                        (OR_SpEmData[pointer]['f1'] != None and
                         OR_SpEmData[pointer]['f1'] >=
                         Bands[table]['SpEmT'][row]['ULfreqStop']) or
                        (OR_SpEmData[pointer]['f2'] != None and
                         OR_SpEmData[pointer]['f2'] >=
                         Bands[table]['SpEmT'][row]['DLfreqStop'])):
                        # last
                        OR_SpEmData.pop(pointer)
                        pointer = len(OR_SpEmData)
                    else:
                        OR_SpEmData.pop(pointer)
                        pointer -= 1
                pointer += 1
            row += 1
        # possible footnote
        Bands[table]['SpEmT_footnote'] = None
        pointer = 0
        while pointer < len(OR_SpEmData):
            if (Bands[table]['ID'] == OR_SpEmData[pointer]['table'] and
                OR_SpEmData[pointer]['row'] == 'footnote'):
                # footnote exists
                Bands[table]['SpEmT_footnote'] = OR_SpEmData[pointer][0]
                OR_SpEmData.pop(pointer)
                pointer = len(OR_SpEmData)
            pointer += 1
print('Override-SpEm not inserted: ', OR_SpEmData) # remains


# Build narrow SpEm tables

def ULDLID(table, row, tag):
    if tag == 'DL':
        not_tag = 'UL'
    else:
        not_tag = 'DL'
    if (Bands[table]['SpEmT'][row]['DLfreqStart'] !=
        Bands[table]['SpEmT'][row]['ULfreqStart'] and
        Bands[table]['SpEmT'][row][not_tag + 'freqStart'] != 'NA'):
        suffix = ' ' + tag
    else:
        suffix = ''
    return Bands[table]['SpEmT'][row]['ID'] + suffix

def SortKeyNarrow(x):
    return (1e10*x['FullBandStart'] - 1e5*x['FullBandStop'] +
            10*ord(x['ID'][0]) + ord(x['ID'][1]) +
            1e-5*x['FreqStart'] - 1e-10*x['FreqStop'])

for table in range(len(Bands)):
    if Bands[table]['SpEmT'] != None:
        Bands[table]['SpEmN'] = []
        for row in range(len(Bands[table]['SpEmT'])):
            if Bands[table]['SpEmT'][row]['ULfreqStart'] != 'NA':
                Bands[table]['SpEmN'].append(
                    {'ID': ULDLID(table, row, 'UL'),
                     'FreqStart': Bands[table]['SpEmT'][row]['ULfreqStart'],
                     'FreqStop': Bands[table]['SpEmT'][row]['ULfreqStop'],
                     'FullBandStart': LookUp(Bands[table]['SpEmT'][row]['ID'],
                                             'ULfreqStart'),
                     'FullBandStop': LookUp(Bands[table]['SpEmT'][row]['ID'],
                                            'ULfreqStop'),
                     'SpEmWA': Bands[table]['SpEmT'][row]['SpEmULWA'],
                     'SpEmAAS': Bands[table]['SpEmT'][row]['SpEmULAAS'],
                     'SpEmMR': Bands[table]['SpEmT'][row]['SpEmULMR'],
                     'SpEmLA': Bands[table]['SpEmT'][row]['SpEmULLA']})
            if (Bands[table]['SpEmT'][row]['DLfreqStart'] != 'NA' and
                Bands[table]['SpEmT'][row]['DLfreqStart'] !=
                Bands[table]['SpEmT'][row]['ULfreqStart']):
                Bands[table]['SpEmN'].append(
                    {'ID': ULDLID(table, row, 'DL'),
                     'FreqStart': Bands[table]['SpEmT'][row]['DLfreqStart'],
                     'FreqStop': Bands[table]['SpEmT'][row]['DLfreqStop'],
                     'FullBandStart': LookUp(Bands[table]['SpEmT'][row]['ID'],
                                             'DLfreqStart'),
                     'FullBandStop': LookUp(Bands[table]['SpEmT'][row]['ID'],
                                            'DLfreqStop'),
                     'SpEmWA': Bands[table]['SpEmT'][row]['SpEmDLWA'],
                     'SpEmAAS': Bands[table]['SpEmT'][row]['SpEmDLAAS'],
                     'SpEmMR': Bands[table]['SpEmT'][row]['SpEmDLMR'],
                     'SpEmLA': Bands[table]['SpEmT'][row]['SpEmDLLA']})
        # Sort narrow SpEm tables
        Bands[table]['SpEmN'].sort(key = SortKeyNarrow)
        # Remove duplicates and zero range lines
        row = 1
        while row < len(Bands[table]['SpEmN']):
            if Bands[table]['SpEmN'][row] == Bands[table]['SpEmN'][row-1] :
                Bands[table]['SpEmN'].pop(row)
            elif (Bands[table]['SpEmN'][row]['FreqStart'] ==
                  Bands[table]['SpEmN'][row]['FreqStop']):
                Bands[table]['SpEmN'].pop(row)
            else:
                row += 1

# Reduce to narrow short list SpEm tables

def SpEmValue(x, opt = 0):
    if (isinstance(x, str) and '/MHz' in x):
        x = x.replace('/MHz', '/1000kHz')
    if (isinstance(x, str) and 'MHz' in x):
        x = x.replace('MHz', '000kHz')
    if (opt == 1 and isinstance(x, str) and
        x[len(x)-9:len(x)] == ' per port' and
        isinstance(SpEmValue(x[0:len(x)-9]) , (int, float)) ):
        return 10 + SpEmValue(x[0:len(x)-9])
    elif isinstance(x, str) and isnumber(x):
        return eval(x)
    elif (isinstance(x, str) and 'kHz' in x and 
        isnumber(x.partition('/')[0]) and
        isnumber(x.partition('/')[2].partition('kHz')[0]) and
        x.partition('/')[2].partition('kHz')[2] == ''):
        return (eval(x.partition('/')[0]) -
                (10*log10(eval(x.partition('/')[2].partition('kHz')[0]))/100) )
    else:
        return x

def SpEm_lt(x, y):
    if (isinstance(x, str) and x[len(x)-9:len(x)] == ' per port' and
        isinstance(y, str) and y[len(y)-9:len(y)] == ' per port' ):
        x = x[0:len(x)-9]
        y = y[0:len(y)-9]
    x = SpEmValue(x)
    y = SpEmValue(y, 1)
    if isinstance(x, (int, float)) and isinstance(y, (int, float)) and x <= y:
        return True
    else:
        return False

def fullbandlist(x):
    x = x.replace('-', ' B').replace('+', ' B').replace('&', ' B') + ' '
    out = []
    while ' ' in x:
        out.append(x.partition(' ')[0])
        x = x.partition(' ')[2]
    for i in range(len(out)):
        while len(out[i]) > 0 and not out[i][-1].isdigit():
            out[i] = out[i][:-1]
    return out

for table in range(len(Bands)):
    if Bands[table]['SpEmT'] != None:
        for row in range(len(Bands[table]['SpEmN'])):
            Bands[table]['SpEmN'][row]['inshort'] = True
            if (Bands[table]['SpEmN'][row]['ID'][:4] == 'Reg ' or
                Bands[table]['SpEmN'][row]['ID'][:3] == 'CR '):
                f1 = Bands[table]['SpEmN'][row]['FreqStart']
                f2 = Bands[table]['SpEmN'][row]['FreqStop']
                for i in range(len(Bands[table]['SpEmN'])):
                    if (i != row and
                        SpEm_lt(Bands[table]['SpEmN'][i]['SpEmWA'],
                                Bands[table]['SpEmN'][row]['SpEmWA']) and
                        SpEm_lt(Bands[table]['SpEmN'][i]['SpEmAAS'],
                                Bands[table]['SpEmN'][row]['SpEmAAS']) and
                        SpEm_lt(Bands[table]['SpEmN'][i]['SpEmMR'],
                                Bands[table]['SpEmN'][row]['SpEmMR']) and
                        SpEm_lt(Bands[table]['SpEmN'][i]['SpEmLA'],
                                Bands[table]['SpEmN'][row]['SpEmLA'])):
                        if (f1 >= Bands[table]['SpEmN'][i]['FreqStart'] and
                            f2 <= Bands[table]['SpEmN'][i]['FreqStop']):
                             Bands[table]['SpEmN'][row]['inshort'] = False
                             for j in range(len(fullbandlist(Bands[table]['ID']))):
                                 if ((fullbandlist(Bands[table]['ID'])[j] + ')') in
                                     Bands[table]['SpEmN'][row]['ID']):
                                     Bands[table]['SpEmN'][row]['inshort'] = True
                                     '''
                                     print('\n', Bands[table]['ID'], '\n',
                                           Bands[table]['SpEmN'][row],
                                           '\n - kept, but covered by:\n',
                                           Bands[table]['SpEmN'][i])
                                     '''
                        elif (f1 >= Bands[table]['SpEmN'][i]['FreqStart'] and
                              f1 <= Bands[table]['SpEmN'][i]['FreqStop']):
                            f1 = Bands[table]['SpEmN'][i]['FreqStop']
                        elif (f2 >= Bands[table]['SpEmN'][i]['FreqStart'] and
                              f2 <= Bands[table]['SpEmN'][i]['FreqStop']):
                            f2 = Bands[table]['SpEmN'][i]['FreqStart']
                            

# ---------------
# Sort and output
# ---------------

Bands.sort(key=lambda x: fillstr(x['ID'], 1))

Table_BandReg = Workbook(write_only = True)
ark = Table_BandReg.create_sheet('BandReg')
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        ark.append([Bands[table]['ID'], Bands[table]['regions']])
Table_BandReg.save('Table_BandReg.xlsx')
print('Table_BandReg.xlsx written')

Table_Blocking = Workbook(write_only = True)
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        tab = Table_Blocking.create_sheet(Bands[table]['ID'] + ' Blocking')
        tab.append(['Potential regions if certified: ' +
                    Bands[table]['regions']])
        if (isinstance(Bands[table]['SensWA'], list) and
            len(Bands[table]['SensWA']) > 1):
            allsame = True
            for i in range(len(Bands[table]['SensWA'])-1):
                if (Bands[table]['SensWA'][i+1][1] !=
                    Bands[table]['SensWA'][0][1]):
                    allsame = False
            if allsame:
                Bands[table]['SensWA'] = Bands[table]['SensWA'][0][1]
        if isinstance(Bands[table]['SensWA'], list):
            out = ''
            for i in range(len(Bands[table]['SensWA'])):
                out += (', ' + Bands[table]['SensWA'][i][0] + ': ' +
                        str(Bands[table]['SensWA'][i][1]))
            Bands[table]['SensWA'] = out[2:]
        label = ['Operating Band / System',
                 'Interferer freq start (MHz)',
                 'Interferer freq stop (MHz)',
                 'Interferer Level WA (dBm)',
                 'Interferer Level AAS (dBm)',
                 'Interferer Level MR (dBm)',
                 'Interferer Level LA (dBm)',
                 'Interferer Type',
                 'Selectivity WA (from LTE5 REFSENS = ' +
                 str(Bands[table]['SensWA']) + ' dBm) (dB)'
                 #, 'Selectivity AAS (dB)',
                 # 'Selectivity MR (dB)', 'Selectivity LA (dB)'
                 ]
        cell = [0]*len(label)
        for col in range(len(label)):
            cell[col] = WriteOnlyCell(tab, value = label[col])
            cell[col].font = Font(bold = True)
            cell[col].fill = PatternFill(fill_type = 'solid',
                                         start_color = 'D9D9D9')
            cell[col].alignment = Alignment(wrap_text = True,
                                            horizontal = 'center',
                                            vertical = 'center')
            cell[col].border = Border(left = Side(border_style='thin',
                                                  color='000000'),
                                      right = Side(border_style='thin',
                                                   color='000000'),
                                      top = Side(border_style='thin',
                                                 color='000000'),
                                      bottom = Side(border_style='thin',
                                                    color='000000'))
        tab.append(cell)            
        for row in range(len(Bands[table]['BlockT'])):
            line = [Bands[table]['BlockT'][row]['ID'],
                    Bands[table]['BlockT'][row]['DLfreqStart'],
                    Bands[table]['BlockT'][row]['DLfreqStop'],
                    Bands[table]['BlockT'][row]['BlockWA'],
                    Bands[table]['BlockT'][row]['BlockAAS'],
                    Bands[table]['BlockT'][row]['BlockMR'],
                    Bands[table]['BlockT'][row]['BlockLA'],
                    Bands[table]['BlockT'][row]['BlockType'],
                    Bands[table]['BlockT'][row]['SelWA']
                    #,
                    #Bands[table]['BlockT'][row]['SelAAS'],
                    #Bands[table]['BlockT'][row]['SelMR'],
                    #Bands[table]['BlockT'][row]['SelLA']
                    ]
            cell = [0]*len(line)
            for col in range(len(line)):
                cell[col] = WriteOnlyCell(tab, value = line[col])
                cell[col].alignment = Alignment(horizontal = 'center',
                                                vertical = 'center')
                cell[col].border = Border(left = Side(border_style='thin',
                                                      color='000000'),
                                          right = Side(border_style='thin',
                                                       color='000000'),
                                          top = Side(border_style='thin',
                                                     color='000000'),
                                          bottom = Side(border_style='thin',
                                                        color='000000'))
            tab.append(cell)
Table_Blocking.save('Table_Blocking_interferer_for_co-location.xlsx')
print('Table_Blocking_interferer_for_co-location.xlsx written')

def head(ULDL, band, case):
    if '-' in band['ID'] and band['MSR'] != None:
        MB = ' per band'
    else:
        MB = ''
    return ("Limit, " + ULDL + " sum over TX ports" +
            MB + ", " + case + " (dBm)")

Table_SpEm = Workbook(write_only=True)
for table in range(len(Bands)):
    if Bands[table]['SpEmT'] != None:
        tab = Table_SpEm.create_sheet(Bands[table]['ID'] + ' SpEm')
        tab.append(['Potential regions if certified: ' +
                    Bands[table]['regions']])
        label = ["Operating Band / System",
                 "UL freq start (MHz)",
                 "UL freq stop (MHz)",
                 head("UL", Bands[table], "WA"),
                 "Limit, UL AAS (dBm)",
                 head("UL", Bands[table], "MR"),
                 head("UL", Bands[table], "LA"),
                 "DL freq start (MHz)",
                 "DL freq stop (MHz)",
                 head("DL", Bands[table], "WA"),
                 "Limit, DL AAS (dBm)",
                 head("DL", Bands[table], "MR"),
                 head("DL", Bands[table], "LA")]
        cell = [0]*len(label)
        for col in range(len(label)):
            cell[col] = WriteOnlyCell(tab, value = label[col] )
            cell[col].font = Font(bold = True)
            cell[col].fill = PatternFill(fill_type = 'solid',
                                         start_color = 'D9D9D9')
            cell[col].alignment = Alignment(wrap_text = True,
                                            horizontal = 'center',
                                            vertical = 'center')
            cell[col].border = Border(left = Side(border_style='thin',
                                                  color='000000'),
                                      right = Side(border_style='thin',
                                                   color='000000'),
                                      top = Side(border_style='thin',
                                                 color='000000'),
                                      bottom = Side(border_style='thin',
                                                    color='000000'))
        tab.append(cell)            
        for row in range(len(Bands[table]['SpEmT'])):
            line = [Bands[table]['SpEmT'][row]['ID'],
                    Bands[table]['SpEmT'][row]['ULfreqStart'],
                    Bands[table]['SpEmT'][row]['ULfreqStop'],
                    Bands[table]['SpEmT'][row]['SpEmULWA'],
                    Bands[table]['SpEmT'][row]['SpEmULAAS'],
                    Bands[table]['SpEmT'][row]['SpEmULMR'],
                    Bands[table]['SpEmT'][row]['SpEmULLA'],
                    Bands[table]['SpEmT'][row]['DLfreqStart'],
                    Bands[table]['SpEmT'][row]['DLfreqStop'],
                    Bands[table]['SpEmT'][row]['SpEmDLWA'],
                    Bands[table]['SpEmT'][row]['SpEmDLAAS'],
                    Bands[table]['SpEmT'][row]['SpEmDLMR'],
                    Bands[table]['SpEmT'][row]['SpEmDLLA']]
            cell = [0]*len(line)
            for col in range(len(line)):
                cell[col] = WriteOnlyCell(tab, value = line[col])
                cell[col].alignment = Alignment(horizontal = 'center',
                                                vertical = 'center')
                cell[col].border = Border(left = Side(border_style='thin',
                                                      color='000000'),
                                          right = Side(border_style='thin',
                                                       color='000000'),
                                          top = Side(border_style='thin',
                                                     color='000000'),
                                          bottom = Side(border_style='thin',
                                                        color='000000'))
            tab.append(cell)
        if Bands[table]['SpEmT_footnote'] != None:
            tab.append([Bands[table]['SpEmT_footnote']])
Table_SpEm.save('Table_SpEm_for_co-location.xlsx')
print('Table_SpEm_for_co-location.xlsx written')

# output narrow SpEm

def HeadNarrow(band, case):
    if '-' in band['ID'] and band['MSR'] != None:
        MB = ' per band'
    else:
        MB = ''
    return ("Limit, sum over TX ports" + MB + ", " + case + " (dBm)")

def WriteNarrow(file, opt = False):
    Table_SpEm_N = Workbook(write_only=True)
    for table in range(len(Bands)):
        if Bands[table]['SpEmT'] != None:
            tab = Table_SpEm_N.create_sheet(Bands[table]['ID'] + ' SpEm')
            tab.append(['Potential regions if certified: ' +
                        Bands[table]['regions']])
            label = ["Operating Band / System", "Freq Start (MHz)",
                     "Freq Stop (MHz)",
                     HeadNarrow(Bands[table], "WA"),
                     "Limit, AAS (dBm)",
                     HeadNarrow(Bands[table], "MR"),
                     HeadNarrow(Bands[table], "LA")]
            cell = [0]*len(label)
            for col in range(len(label)):
                cell[col] = WriteOnlyCell(tab, value = label[col] )
                cell[col].font = Font(bold = True)
                cell[col].fill = PatternFill(fill_type = 'solid',
                                             start_color = 'D9D9D9')
                cell[col].alignment = Alignment(wrap_text = True,
                                                horizontal = 'center',
                                                vertical = 'center')
                cell[col].border = Border(left = Side(border_style='thin',
                                                      color='000000'),
                                          right = Side(border_style='thin',
                                                       color='000000'),
                                          top = Side(border_style='thin',
                                                     color='000000'),
                                          bottom = Side(border_style='thin',
                                                        color='000000'))
            tab.append(cell)            
            for row in range(len(Bands[table]['SpEmN'])):
                line = [Bands[table]['SpEmN'][row]['ID'],
                        Bands[table]['SpEmN'][row]['FreqStart'],
                        Bands[table]['SpEmN'][row]['FreqStop'],
                        Bands[table]['SpEmN'][row]['SpEmWA'],
                        Bands[table]['SpEmN'][row]['SpEmAAS'],
                        Bands[table]['SpEmN'][row]['SpEmMR'],
                        Bands[table]['SpEmN'][row]['SpEmLA']]
                cell = [0]*len(line)
                for col in range(len(line)):
                    cell[col] = WriteOnlyCell(tab, value = line[col])
                    cell[col].alignment = Alignment(horizontal = 'center',
                                                    vertical = 'center')
                    cell[col].border = Border(left = Side(border_style='thin',
                                                          color='000000'),
                                              right = Side(border_style='thin',
                                                           color='000000'),
                                              top = Side(border_style='thin',
                                                         color='000000'),
                                              bottom =Side(border_style='thin',
                                                           color='000000'))
                if ((opt and Bands[table]['SpEmN'][row]['inshort']) or
                    not opt):
                    tab.append(cell)
            if Bands[table]['SpEmT_footnote'] != None:
                tab.append([Bands[table]['SpEmT_footnote']])
    Table_SpEm_N.save(file)

# WriteNarrow('Table_SpEm_for_co-location_narrow.xlsx')
# print('Table_SpEm_for_co-location_narrow.xlsx written')

WriteNarrow('Table_SpEm_for_co-location_narrow_short.xlsx', True)
print('Table_SpEm_for_co-location_narrow_short.xlsx written')

# ----------------------------------
# Generate PrelAlloc blocking tables
# ----------------------------------

# Add general blocking
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        Bands[table]['BlockT'].append({'ID': '',
                                       'DLfreqStart': 1,
                                       'DLfreqStop': 12750,
                                       'BlockWA': '-15 @ 6 dB SD',
                                       'BlockAAS': '-15 @ 6 dB SD',
                                       'BlockMR': '-15 @ 6 dB SD',
                                       'BlockLA': '-15 @ 6 dB SD',
                                       'BlockType': 'CW'})

# Merge adjacent or overlapping bands with identical req:
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        while True:
            change = 0
            row = 0
            while row < len(Bands[table]['BlockT']):
                line = 0
                while line < len(Bands[table]['BlockT']):
                    if (line != row and
                        Bands[table]['BlockT'][row]['BlockWA'] ==
                        Bands[table]['BlockT'][line]['BlockWA'] and
                        Bands[table]['BlockT'][row]['BlockAAS'] ==
                        Bands[table]['BlockT'][line]['BlockAAS'] and
                        Bands[table]['BlockT'][row]['BlockMR'] ==
                        Bands[table]['BlockT'][line]['BlockMR'] and
                        Bands[table]['BlockT'][row]['BlockLA'] ==
                        Bands[table]['BlockT'][line]['BlockLA'] and
                        Bands[table]['BlockT'][line]['BlockType'] in
                        Bands[table]['BlockT'][row]['BlockType']):
                        # bands have same levels
                        if ((Bands[table]['BlockT'][row]['DLfreqStart'] <=
                             Bands[table]['BlockT'][line]['DLfreqStart'] and
                             Bands[table]['BlockT'][row]['DLfreqStop'] >=
                             Bands[table]['BlockT'][line]['DLfreqStop'] ) or
                            # row cover line
                            (Bands[table]['BlockT'][row]['DLfreqStart'] >=
                             Bands[table]['BlockT'][line]['DLfreqStart'] and
                             Bands[table]['BlockT'][row]['DLfreqStop'] <=
                             Bands[table]['BlockT'][line]['DLfreqStop'] ) or
                            # row inside line
                            (Bands[table]['BlockT'][row]['DLfreqStart'] >=
                             Bands[table]['BlockT'][line]['DLfreqStart'] and
                             Bands[table]['BlockT'][row]['DLfreqStart'] <=
                             Bands[table]['BlockT'][line]['DLfreqStop'] ) or
                            # line below row
                            (Bands[table]['BlockT'][row]['DLfreqStop'] >=
                             Bands[table]['BlockT'][line]['DLfreqStart'] and
                             Bands[table]['BlockT'][row]['DLfreqStop'] <=
                             Bands[table]['BlockT'][line]['DLfreqStop'] )):
                            # line above row
                            # bands adjacent or overlapping!
                            Bands[table]['BlockT'][row]['ID'] += \
                            '; ' + Bands[table]['BlockT'][line]['ID']
                            Bands[table]['BlockT'][row]['DLfreqStart'] = \
                            min(Bands[table]['BlockT'][row]['DLfreqStart'],
                                Bands[table]['BlockT'][line]['DLfreqStart'])
                            Bands[table]['BlockT'][row]['DLfreqStop'] = \
                            max(Bands[table]['BlockT'][row]['DLfreqStop'],
                                Bands[table]['BlockT'][line]['DLfreqStop'])
                            Bands[table]['BlockT'].pop(line)
                            if row > line:
                                row -= 1
                            if line != 0:
                                line -= 1
                            change = 1
                    line += 1
                row += 1
            if change == 0:
                break

def blockvalue(x, y = 1):
    if (isinstance(x, str) and
        isnumber(x.partition(' @ ')[0]) and
        isnumber(x.partition(' @ ')[2].partition(' dB SD')[0]) and
        x.partition(' @ ')[2].partition(' dB SD')[2] == ''):
        return (eval(x.partition(' @ ')[0]) -
                10*log10(10**(eval(x.partition(' @ ')[2].
                                   partition(' dB SD')[0])/10)-1))
    elif isinstance(x, (int, float)):
        return x - 10*log10(10**(y/10)-1)
    else:
        return x

def block_gt(x, x_type, y, y_type):
    if ((isinstance(x, (int, float)) and isinstance(y, (int, float))) and
        ((x >= y and y_type in x_type) or
         (x >= y - 10 and y_type == 'CW'))):
        return True
    else:
        return False

for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        while True:
            change = 0
            row = 0
            while row < len(Bands[table]['BlockT']):
                line = 0
                while line < len(Bands[table]['BlockT']):
                    if (line != row and
                        block_gt(blockvalue(Bands[table]['BlockT'][row]\
                                            ['BlockWA'], 0.4),
                                 Bands[table]['BlockT'][row]['BlockType'], 
                                 blockvalue(Bands[table]['BlockT'][line]\
                                            ['BlockWA'], 0.4),
                                 Bands[table]['BlockT'][line]['BlockType']) and
                        block_gt(blockvalue(Bands[table]['BlockT'][row]\
                                            ['BlockAAS']),
                                 Bands[table]['BlockT'][row]['BlockType'], 
                                 blockvalue(Bands[table]['BlockT'][line]\
                                            ['BlockAAS']),
                                 Bands[table]['BlockT'][line]['BlockType']) and
                        block_gt(blockvalue(Bands[table]['BlockT'][row]\
                                            ['BlockMR']),
                                 Bands[table]['BlockT'][row]['BlockType'], 
                                 blockvalue(Bands[table]['BlockT'][line]\
                                            ['BlockMR']),
                                 Bands[table]['BlockT'][line]['BlockType']) and
                        block_gt(blockvalue(Bands[table]['BlockT'][row]\
                                            ['BlockLA']),
                                 Bands[table]['BlockT'][row]['BlockType'], 
                                 blockvalue(Bands[table]['BlockT'][line]\
                                            ['BlockLA']),
                                 Bands[table]['BlockT'][line]['BlockType'])):
                        # band @ row >= band @ line
                        if (Bands[table]['BlockT'][row]['DLfreqStart'] <=
                            Bands[table]['BlockT'][line]['DLfreqStart'] and
                            Bands[table]['BlockT'][row]['DLfreqStop'] >=
                            Bands[table]['BlockT'][line]['DLfreqStop']):
                            # row cover line
                            Bands[table]['BlockT'][row]['ID'] += \
                            '; ' + Bands[table]['BlockT'][line]['ID']
                            Bands[table]['BlockT'].pop(line)
                            if row > line:
                                row -= 1
                            if line != 0:
                                line -= 1
                            change = 1
                        elif (Bands[table]['BlockT'][row]['DLfreqStart'] >
                              Bands[table]['BlockT'][line]['DLfreqStart'] and
                              Bands[table]['BlockT'][row]['DLfreqStop'] <
                              Bands[table]['BlockT'][line]['DLfreqStop']):
                            # row inside line
                            if Bands[table]['BlockT'][line]['ID'] != '':
                                Bands[table]['BlockT'][row]['ID'] += \
                                '; ' + Bands[table]['BlockT'][line]['ID']
                            Bands[table]['BlockT'].append(Bands[table]
                                                          ['BlockT']
                                                          [line].copy())
                            Bands[table]['BlockT'][line]['DLfreqStop'] = \
                            Bands[table]['BlockT'][row]['DLfreqStart']
                            Bands[table]['BlockT'][len(Bands[table]
                                                       ['BlockT']) - 1]\
                                                       ['DLfreqStart'] = \
                            Bands[table]['BlockT'][row]['DLfreqStop']
                            change = 1
                        elif (Bands[table]['BlockT'][row]['DLfreqStart'] >
                              Bands[table]['BlockT'][line]['DLfreqStart'] and
                              Bands[table]['BlockT'][row]['DLfreqStart'] <
                              Bands[table]['BlockT'][line]['DLfreqStop'] and
                              Bands[table]['BlockT'][row]['DLfreqStop'] >=
                              Bands[table]['BlockT'][line]['DLfreqStop']):
                            # line below row
                            Bands[table]['BlockT'][row]['ID'] += \
                            '; ' + Bands[table]['BlockT'][line]['ID']
                            Bands[table]['BlockT'][line]['DLfreqStop'] = \
                            Bands[table]['BlockT'][row]['DLfreqStart']
                            change = 1
                        elif (Bands[table]['BlockT'][row]['DLfreqStop'] >
                              Bands[table]['BlockT'][line]['DLfreqStart'] and
                              Bands[table]['BlockT'][row]['DLfreqStop'] <
                              Bands[table]['BlockT'][line]['DLfreqStop'] and
                              Bands[table]['BlockT'][row]['DLfreqStart'] <=
                              Bands[table]['BlockT'][line]['DLfreqStart']):
                            # line above row
                            Bands[table]['BlockT'][row]['ID'] += \
                            '; ' + Bands[table]['BlockT'][line]['ID']
                            Bands[table]['BlockT'][line]['DLfreqStart'] = \
                            Bands[table]['BlockT'][row]['DLfreqStop']
                            change = 1
                    line += 1
                row += 1
            if change == 0:
                break

Compact_Blocking = Workbook(write_only=True)
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        # sort blocking tables
        Bands[table]['BlockT'].sort(key = sortkey)
        tab = Compact_Blocking.create_sheet(Bands[table]['ID'])
        label = ["Operating Bands / Systems",
                 "Interferer freq start (MHz)",
                 "Interferer freq stop (MHz)",
                 "Interferer Level WA (dBm)",
                 "Interferer Level AAS (dBm)",
                 "Interferer Level MR (dBm)",
                 "Interferer Level LA (dBm)",
                 "Interferer Type"]
        cell = [0]*len(label)
        for col in range(len(label)):
            cell[col] = WriteOnlyCell(tab, value = label[col])
        tab.append(cell)            
        for row in range(len(Bands[table]['BlockT'])):
            line = [Bands[table]['BlockT'][row]['ID'],
                    Bands[table]['BlockT'][row]['DLfreqStart'],
                    Bands[table]['BlockT'][row]['DLfreqStop'],
                    Bands[table]['BlockT'][row]['BlockWA'],
                    Bands[table]['BlockT'][row]['BlockAAS'],
                    Bands[table]['BlockT'][row]['BlockMR'],
                    Bands[table]['BlockT'][row]['BlockLA'],
                    Bands[table]['BlockT'][row]['BlockType']]
            cell = [0]*len(line)
            for col in range(len(line)):
                cell[col] = WriteOnlyCell(tab, value = line[col])
            tab.append(cell)
Compact_Blocking.save('Table_Blocking_PrelAlloc.xlsx')
print('Table_Blocking_PrelAlloc.xlsx written')

# ------------------------------
# Generate PrelAlloc SpEm tables
# ------------------------------

def zero_if_not_number(x):
    if isnumber(x):
        return x
    else:
        return 0


# Add general SpEm
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        Bands[table]['SpEmN'].append({'ID': '',
                                      'FreqStart': 0.009,
                                      'FreqStop': 0.15,
                                      'SpEmWA': '-36/kHz per port',
                                      'SpEmAAS': '-36/kHz',
                                      'SpEmMR': '-36/kHz per port',
                                      'SpEmLA': '-36/kHz per port'})
        Bands[table]['SpEmN'].append({'ID': '',
                                      'FreqStart': 0.15,
                                      'FreqStop': 30,
                                      'SpEmWA': '-36/10kHz per port',
                                      'SpEmAAS': '-36/10kHz',
                                      'SpEmMR': '-36/10kHz per port',
                                      'SpEmLA': '-36/10kHz per port'})
        Bands[table]['SpEmN'].append({'ID': '',
                                      'FreqStart': 30,
                                      'FreqStop': 1000,
                                      'SpEmWA': '-36 per port',
                                      'SpEmAAS': -36,
                                      'SpEmMR': '-36 per port',
                                      'SpEmLA': '-36 per port'})
        Bands[table]['SpEmN'].append({'ID': '',
                                      'FreqStart': 1000,
                                      'FreqStop': max(12750,
                                                      zero_if_not_number(\
                                      5*Bands[table]['DLfreqStop'])),
                                      'SpEmWA': '-30/MHz per port',
                                      'SpEmAAS': '-30/MHz',
                                      'SpEmMR': '-30/MHz per port',
                                      'SpEmLA': '-30/MHz per port'})

# Additional SpEm for BC2 (only works for single bands)
for table in range(len(Bands)):
    for b in ['0', '2', '3', '5', '8']:
        if (('B000' + b) in mbfill(Bands[table]['ID']) and
            'M' not in str(Bands[table]['MSR'])
            and 'T' in str(Bands[table]['MSR'])):
            if Bands[table]['DLfreqStart'] > 1030:
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart': 500,
                                              'FreqStop': 1000,
                                              'SpEmWA': '-36/3MHz per port',
                                              'SpEmAAS': '-36/3MHz',
                                              'SpEmMR': '-36/3MHz per port',
                                              'SpEmLA': '-36/3MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart': 1000,
                                              'FreqStop':
                                              Bands[table]['DLfreqStart'] - 30,
                                              'SpEmWA': '-30/3MHz per port',
                                              'SpEmAAS': '-30/3MHz',
                                              'SpEmMR': '-30/3MHz per port',
                                              'SpEmLA': '-30/3MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStop'] + 30,
                                              'FreqStop': 12750,
                                              'SpEmWA': '-30/3MHz per port',
                                              'SpEmAAS': '-30/3MHz',
                                              'SpEmMR': '-30/3MHz per port',
                                              'SpEmLA': '-30/3MHz per port'})
            else:
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart': 500,
                                              'FreqStop':
                                              Bands[table]['DLfreqStart'] - 30,
                                              'SpEmWA': '-36/3MHz per port',
                                              'SpEmAAS': '-36/3MHz',
                                              'SpEmMR': '-36/3MHz per port',
                                              'SpEmLA': '-36/3MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStart'] - 30,
                                              'FreqStop':
                                              Bands[table]['DLfreqStart'] - 20,
                                              'SpEmWA': '-36/MHz per port',
                                              'SpEmAAS': '-36/MHz',
                                              'SpEmMR': '-36/MHz per port',
                                              'SpEmLA': '-36/MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStart'] - 20,
                                              'FreqStop':
                                              Bands[table]['DLfreqStart'] - 10,
                                              'SpEmWA': '-36/300kHz per port',
                                              'SpEmAAS': '-36/300kHz',
                                              'SpEmMR': '-36/300kHz per port',
                                              'SpEmLA': '-36/300kHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStop'] + 10,
                                              'FreqStop':
                                              Bands[table]['DLfreqStop'] + 20,
                                              'SpEmWA': '-36/300kHz per port',
                                              'SpEmAAS': '-36/300kHz',
                                              'SpEmMR': '-36/300kHz per port',
                                              'SpEmLA': '-36/300kHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStop'] + 20,
                                              'FreqStop':
                                              Bands[table]['DLfreqStop'] + 30,
                                              'SpEmWA': '-36/MHz per port',
                                              'SpEmAAS': '-36/MHz',
                                              'SpEmMR': '-36/MHz per port',
                                              'SpEmLA': '-36/MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart':
                                              Bands[table]['DLfreqStop'] + 30,
                                              'FreqStop': 1000,
                                              'SpEmWA': '-36/3MHz per port',
                                              'SpEmAAS': '-36/3MHz',
                                              'SpEmMR': '-36/3MHz per port',
                                              'SpEmLA': '-36/3MHz per port'})
                Bands[table]['SpEmN'].append({'ID': '',
                                              'FreqStart': 1000,
                                              'FreqStop': 12750,
                                              'SpEmWA': '-30/3MHz per port',
                                              'SpEmAAS': '-30/3MHz',
                                              'SpEmMR': '-30/3MHz per port',
                                              'SpEmLA': '-30/3MHz per port'})

# Merge adjacent or overlapping bands with identical req:
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        while True:
            change = 0
            row = 0
            while row < len(Bands[table]['SpEmN']):
                line = 0
                while line < len(Bands[table]['SpEmN']):
                    if (line != row and
                        Bands[table]['SpEmN'][row]['SpEmWA'] ==
                        Bands[table]['SpEmN'][line]['SpEmWA'] and
                        Bands[table]['SpEmN'][row]['SpEmAAS'] ==
                        Bands[table]['SpEmN'][line]['SpEmAAS'] and
                        Bands[table]['SpEmN'][row]['SpEmMR'] ==
                        Bands[table]['SpEmN'][line]['SpEmMR'] and
                        Bands[table]['SpEmN'][row]['SpEmLA'] ==
                        Bands[table]['SpEmN'][line]['SpEmLA']):
                        # bands have same levels
                        if ((Bands[table]['SpEmN'][row]['FreqStart'] <=
                             Bands[table]['SpEmN'][line]['FreqStart'] and
                             Bands[table]['SpEmN'][row]['FreqStop'] >=
                             Bands[table]['SpEmN'][line]['FreqStop'] ) or
                            # row cover line
                            (Bands[table]['SpEmN'][row]['FreqStart'] >=
                             Bands[table]['SpEmN'][line]['FreqStart'] and
                             Bands[table]['SpEmN'][row]['FreqStop'] <=
                             Bands[table]['SpEmN'][line]['FreqStop'] ) or
                            # row inside line
                            (Bands[table]['SpEmN'][row]['FreqStart'] >=
                             Bands[table]['SpEmN'][line]['FreqStart'] and
                             Bands[table]['SpEmN'][row]['FreqStart'] <=
                             Bands[table]['SpEmN'][line]['FreqStop'] ) or
                            # line below row
                            (Bands[table]['SpEmN'][row]['FreqStop'] >=
                             Bands[table]['SpEmN'][line]['FreqStart'] and
                             Bands[table]['SpEmN'][row]['FreqStop'] <=
                             Bands[table]['SpEmN'][line]['FreqStop'] )):
                            # line above row
                            # bands adjacent or overlapping!
                            Bands[table]['SpEmN'][row]['ID'] += \
                            '; ' + Bands[table]['SpEmN'][line]['ID']
                            Bands[table]['SpEmN'][row]['FreqStart'] = \
                            min(Bands[table]['SpEmN'][row]['FreqStart'],
                                Bands[table]['SpEmN'][line]['FreqStart'])
                            Bands[table]['SpEmN'][row]['FreqStop'] = \
                            max(Bands[table]['SpEmN'][row]['FreqStop'],
                                Bands[table]['SpEmN'][line]['FreqStop'])
                            Bands[table]['SpEmN'].pop(line)
                            if row > line:
                                row -= 1
                            if line != 0:
                                line -= 1
                            change = 1
                    line += 1
                row += 1
            if change == 0:
                break

def SpEm_lt(x, y):
    if (isinstance(x, str) and x[len(x)-9:len(x)] == ' per port' and
        isinstance(y, str) and y[len(y)-9:len(y)] == ' per port' ):
        x = x[0:len(x)-9]
        y = y[0:len(y)-9]
    x = SpEmValue(x)
    y = SpEmValue(y, 1)
    if isinstance(x, (int, float)) and isinstance(y, (int, float)) and x <= y:
        return True
    else:
        return False

for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        while True:
            change = 0
            row = 0
            while row < len(Bands[table]['SpEmN']):
                line = 0
                while line < len(Bands[table]['SpEmN']):
                    if (line != row and
                        SpEm_lt(Bands[table]['SpEmN'][row]['SpEmWA'],
                                Bands[table]['SpEmN'][line]['SpEmWA']) and
                        SpEm_lt(Bands[table]['SpEmN'][row]['SpEmAAS'],
                                Bands[table]['SpEmN'][line]['SpEmAAS']) and
                        SpEm_lt(Bands[table]['SpEmN'][row]['SpEmMR'],
                                Bands[table]['SpEmN'][line]['SpEmMR']) and
                        SpEm_lt(Bands[table]['SpEmN'][row]['SpEmLA'],
                                Bands[table]['SpEmN'][line]['SpEmLA']) ):
                        # band @ row <= band @ line
                        if (Bands[table]['SpEmN'][row]['FreqStart'] <=
                            Bands[table]['SpEmN'][line]['FreqStart'] and
                            Bands[table]['SpEmN'][row]['FreqStop'] >=
                            Bands[table]['SpEmN'][line]['FreqStop']):
                            # row cover line
                            Bands[table]['SpEmN'][row]['ID'] += \
                            '; ' + Bands[table]['SpEmN'][line]['ID']
                            Bands[table]['SpEmN'].pop(line)
                            if row > line:
                                row -= 1
                            if line != 0:
                                line -= 1
                            change = 1
                        elif (Bands[table]['SpEmN'][row]['FreqStart'] >
                              Bands[table]['SpEmN'][line]['FreqStart'] and
                              Bands[table]['SpEmN'][row]['FreqStop'] <
                              Bands[table]['SpEmN'][line]['FreqStop']):
                            # row inside line
                            if Bands[table]['SpEmN'][line]['ID'] != '':
                                Bands[table]['SpEmN'][row]['ID'] += \
                                '; ' + Bands[table]['SpEmN'][line]['ID']
                            Bands[table]['SpEmN'].append(Bands[table]['SpEmN']
                                                         [line].copy())
                            Bands[table]['SpEmN'][line]['FreqStop'] = \
                            Bands[table]['SpEmN'][row]['FreqStart']
                            Bands[table]['SpEmN']\
                            [len(Bands[table]['SpEmN']) - 1]['FreqStart'] = \
                            Bands[table]['SpEmN'][row]['FreqStop']
                            change = 1
                        elif (Bands[table]['SpEmN'][row]['FreqStart'] >
                              Bands[table]['SpEmN'][line]['FreqStart'] and
                              Bands[table]['SpEmN'][row]['FreqStart'] <
                              Bands[table]['SpEmN'][line]['FreqStop'] and
                              Bands[table]['SpEmN'][row]['FreqStop'] >=
                              Bands[table]['SpEmN'][line]['FreqStop']):
                            # line below row
                            Bands[table]['SpEmN'][row]['ID'] += \
                            '; ' + Bands[table]['SpEmN'][line]['ID']
                            Bands[table]['SpEmN'][line]['FreqStop'] = \
                            Bands[table]['SpEmN'][row]['FreqStart']
                            change = 1
                        elif (Bands[table]['SpEmN'][row]['FreqStop'] >
                              Bands[table]['SpEmN'][line]['FreqStart'] and
                              Bands[table]['SpEmN'][row]['FreqStop'] <
                              Bands[table]['SpEmN'][line]['FreqStop'] and
                              Bands[table]['SpEmN'][row]['FreqStart'] <=
                              Bands[table]['SpEmN'][line]['FreqStart']):
                            # line above row
                            Bands[table]['SpEmN'][row]['ID'] += \
                            '; ' + Bands[table]['SpEmN'][line]['ID']
                            Bands[table]['SpEmN'][line]['FreqStart'] = \
                            Bands[table]['SpEmN'][row]['FreqStop']
                            change = 1
                    line += 1
                row += 1
            if change == 0:
                break

def sortkey2(x):
    return 1e6*x['FreqStart'] - x['FreqStop']

Compact_SpEm = Workbook(write_only=True)
for table in range(len(Bands)):
    if Bands[table]['BlockT'] != None:
        # sort blocking tables
        Bands[table]['SpEmN'].sort(key = sortkey2)
        tab = Compact_SpEm.create_sheet(Bands[table]['ID'])
        label = ["Operating Bands / Systems",
                 "freq start (MHz)",
                 "freq stop (MHz)",
                 "SpEm WA (dBm)",
                 "SpEm AAS (dBm)",
                 "SpEm MR (dBm)",
                 "SpEm LA (dBm)"]
        cell = [0]*len(label)
        for col in range(len(label)):
            cell[col] = WriteOnlyCell(tab, value = label[col])
        tab.append(cell)            
        for row in range(len(Bands[table]['SpEmN'])):
            line = [Bands[table]['SpEmN'][row]['ID'],
                    Bands[table]['SpEmN'][row]['FreqStart'],
                    Bands[table]['SpEmN'][row]['FreqStop'],
                    Bands[table]['SpEmN'][row]['SpEmWA'],
                    Bands[table]['SpEmN'][row]['SpEmAAS'],
                    Bands[table]['SpEmN'][row]['SpEmMR'],
                    Bands[table]['SpEmN'][row]['SpEmLA']]
            cell = [0]*len(line)
            for col in range(len(line)):
                cell[col] = WriteOnlyCell(tab, value = line[col])
            tab.append(cell)
Compact_SpEm.save('Table_SpEm_PrelAlloc.xlsx')
print('Table_SpEm_PrelAlloc.xlsx written')

# --------------
# Ericsson Bands
# --------------

def ComCheck(x):
    if x != None:
        if x[0] != '_':
            return x
        else:
            return None
    else:
        return None

def RegCheck(x):
    if x != 'none.':
        return x.replace('GSM_51.021', '').lstrip(' ,.')
    else:
        return None

def nr(x):
    if x != 'BH':
        return int(x.lstrip('B')[:CountDig(x)])
    else:
        return 'H'

Table_Bands = Workbook(write_only = True)

FDD = Table_Bands.create_sheet('FDD')
FDD.append(['Ericsson\nBand', 'UL\n(MHz)', 'DL\n(MHz)', 'Comment',
            'Potential regions if certified'])

SDL = Table_Bands.create_sheet('SDL')
SDL.append(['Ericsson\nBand', 'DL\n(MHz)', 'Comment',
            'Potential regions if certified'])

SUL = Table_Bands.create_sheet('SUL')
SUL.append(['Ericsson\nBand', 'UL\n(MHz)', 'Comment',
            'Potential regions if certified'])

TDD = Table_Bands.create_sheet('TDD')
TDD.append(['Ericsson\nBand', 'UL, DL\n(MHz)', 'Comment',
            'Potential regions if certified'])

for row in range(len(Bands)):
    if 'L' in str(Bands[row]['MSR']):
        if Bands[row]['type'] == 'FDD':
            FDD.append([Bands[row]['ID'].replace('_', ' ').replace('G1.5',
                                                                   '2219'),
                        str(Bands[row]['ULfreqStart']) + '-' +
                        str(Bands[row]['ULfreqStop']),
                        str(Bands[row]['DLfreqStart']) + '-' +
                        str(Bands[row]['DLfreqStop']),
                        ComCheck(Bands[row]['Comment']),
                        RegCheck(Bands[row]['regions']),
                        nr(Bands[row]['ID'])])
        elif Bands[row]['type'] == 'SDL':
            SDL.append([Bands[row]['ID'].replace('_', ' ').replace('G1.5',
                                                                   '2219'),
                        str(Bands[row]['DLfreqStart']) + '-' +
                        str(Bands[row]['DLfreqStop']),
                        ComCheck(Bands[row]['Comment']),
                        RegCheck(Bands[row]['regions']),
                        nr(Bands[row]['ID'])])
        elif Bands[row]['type'] == 'SUL':
            SUL.append([Bands[row]['ID'].replace('_', ' ').replace('G1.5',
                                                                   '2219'),
                        str(Bands[row]['ULfreqStart']) + '-' +
                        str(Bands[row]['ULfreqStop']),
                        ComCheck(Bands[row]['Comment']),
                        RegCheck(Bands[row]['regions']),
                        nr(Bands[row]['ID'])])
        elif Bands[row]['type'] == 'TDD':
            TDD.append([Bands[row]['ID'].replace('_', ' ').replace('G1.5',
                                                                   '2219'),
                        str(Bands[row]['ULfreqStart']) + '-' +
                        str(Bands[row]['ULfreqStop']),
                        ComCheck(Bands[row]['Comment']),
                        RegCheck(Bands[row]['regions']),
                        nr(Bands[row]['ID'])])
Table_Bands.save('Table_Ericsson_Bands.xlsx')
print('Table_Ericsson_Bands.xlsx written')

if OR_SpEmData != [] or OR_BlockData != []:
    input('Press enter to exit')
